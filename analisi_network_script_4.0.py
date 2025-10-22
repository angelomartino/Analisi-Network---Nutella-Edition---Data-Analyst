import os
import sys
import csv
import ipaddress
import re
import threading
import datetime
import time
import queue
import tkinter as tk
from PIL import Image, ImageTk
from tkinter import ttk, filedialog, messagebox
from typing import Iterable
from tempfile import NamedTemporaryFile
from pathlib import Path
try:
    from openpyxl import load_workbook  # lettura Excel senza dipendere da pandas
except Exception:
    load_workbook = None  # gestisco più avanti se il modulo non è installato


# ========================================================================
# CONFIGURATION & DATA STRUCTURES
# ========================================================================

# Location normalization mapping
LOC_NORMALIZATION = {
    "risc-unknown-internet": "Public",
    "risc-unknown-private": "Private",
}

# Port-based comment rules
PORT_COMMENTS = {
    25: "SMTP",
    8080: "Proxy",
    10065: "Zscaler",
    383: "Rule OMI",
}

# Esempi
""" Service-to-IP mapping
SERVIZIO_TO_IPS = {
    'Azure Migrate': ['10.11.12.13', '10.11.12.13', '10.11.12.13'],
    'Flexera': ['10.11.12.13', '10.11.12.13'],
    'KMS': ['10.11.12.13', '10.11.12.13'],
    'LoadBalancer': ['10.11.12.13'],
    'Monitoring': ['10.11.12.13', '10.11.12.13', '10.11.12.13', '10.11.12.13'],
    'Proxy Germania': ['10.11.12.13'],
    'Proxy Italia': ['10.11.12.13'],
}"""

"""Location-to-subnet mapping
LOCATION_TO_SUBNETS = {
    'ALBA_AUT NET': ['10.11.12.13/16', '10.11.12.13/20'],
    'ALBA_DMZ EXT': ['10.11.12.13/24', '10.11.12.13/27'],
    'ALBA_DMZ INT': ['10.11.12.13/24', '10.11.12.13/25', '10.11.12.13/28', '10.11.12.13/24', '10.11.12.13/27'],
}"""

SERVIZIO_TO_IPS: dict = {}
LOCATION_TO_SUBNETS: dict = {}


# ========================================================================
# HELPER FUNCTIONS - IP & NETWORK UTILITIES
# ========================================================================

def build_ip_to_service_index(service_map: dict) -> dict:

    # Build reverse index from IP to service name for O(1) lookups.

    index = {}
    for service, ips in service_map.items():
        for ip in ips:
            index[str(ip).strip()] = service
    return index


def build_subnet_index(subnets_by_location: dict) -> list:

    # Parse and sort subnets in  "Location-to-subnet" mapping for efficient longest-prefix matching.
    
    compiled = []
    for location, cidrs in subnets_by_location.items():
        for cidr in cidrs:
            try:
                network = ipaddress.ip_network(str(cidr).strip(), strict=False)
                compiled.append((network, location)) #una lista di tuple
            except ValueError:
                continue  # Skip invalid CIDR notation
    
    # Sort by prefix length (most specific first) for longest-prefix match
    compiled.sort(key=lambda item: item[0].prefixlen, reverse=True)
    return compiled


def find_location_for_ip(ip_str: str, subnet_index: list) -> str | None:

    # Find location for an IP using longest-prefix match.
    
    ip_str = (ip_str or "").strip()
    if not ip_str:
        return None
    
    try:
        ip_obj = ipaddress.IPv4Address(ip_str)
    except ValueError:
        return None
    
    # Return first match (most specific subnet)
    for network, location in subnet_index:
        if ip_obj in network:
            return location
    
    return None


def port_matches(target_port: int, port_value) -> bool:
    """
    Check if target port number appears as a whole number in the port value.
    
    Args:
        target_port: Port number to search for
        port_value: Value from CSV (may contain multiple ports)
        
    Returns:
        True if target port is found as a complete number
    """
    port_str = str(port_value or "")
    # Use word boundary regex to match whole numbers only
    return bool(re.search(rf'(?<!\d){target_port}(?!\d)', port_str))


# >>> NEW  (mettila dopo la sezione "HELPER FUNCTIONS - IP & NETWORK UTILITIES")

def load_mappings_from_excel(xlsx_path: str) -> tuple[dict, dict]:
    """
    Legge un file Excel con due fogli:
      - 'location' con colonne: location, subnet
      - 'servizi'  con colonne: servizio, hostname, ip_address (si usano solo servizio, ip_address)

    Ritorna:
      (servizio_to_ips: dict[str, list[str]], location_to_subnets: dict[str, list[str]])
    """
    if not xlsx_path:
        return {}, {}

    if load_workbook is None:
        raise RuntimeError("Per leggere l'Excel serve 'openpyxl'. Installa con: pip install openpyxl")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # --- Foglio location ---
    if "location" not in wb.sheetnames:
        raise ValueError("Nel file Excel manca il foglio 'location'.")

    ws_loc = wb["location"]
    # Trovo gli indici delle colonne (case-insensitive)
    headers_loc = {str((ws_loc.cell(row=1, column=c).value or "")).strip().lower(): c
                   for c in range(1, ws_loc.max_column + 1)}
    for needed in ("location", "subnet"):
        if needed not in headers_loc:
            raise ValueError(f"Nel foglio 'location' manca la colonna '{needed}'.")

    col_loc = headers_loc["location"]
    col_sub = headers_loc["subnet"]

    location_to_subnets: dict[str, list[str]] = {}
    for r in range(2, ws_loc.max_row + 1):
        loc = str(ws_loc.cell(row=r, column=col_loc).value or "").strip()
        sub = str(ws_loc.cell(row=r, column=col_sub).value or "").strip()
        if not loc or not sub:
            continue
        # valida sintassi CIDR; se non valida, skip silenzioso
        try:
            ipaddress.ip_network(sub, strict=False)
        except Exception:
            continue
        location_to_subnets.setdefault(loc, [])
        if sub not in location_to_subnets[loc]:
            location_to_subnets[loc].append(sub)

    # --- Foglio servizi ---
    if "servizi" not in wb.sheetnames:
        raise ValueError("Nel file Excel manca il foglio 'servizi'.")

    ws_srv = wb["servizi"]
    headers_srv = {str((ws_srv.cell(row=1, column=c).value or "")).strip().lower(): c
                   for c in range(1, ws_srv.max_column + 1)}
    for needed in ("servizio", "ip_address"):
        if needed not in headers_srv:
            raise ValueError(f"Nel foglio 'servizi' manca la colonna '{needed}'.")

    col_srv = headers_srv["servizio"]
    col_ip  = headers_srv["ip_address"]

    servizio_to_ips: dict[str, list[str]] = {}
    for r in range(2, ws_srv.max_row + 1):
        srv = str(ws_srv.cell(row=r, column=col_srv).value or "").strip()
        ip  = str(ws_srv.cell(row=r, column=col_ip).value or "").strip()
        if not srv or not ip:
            continue
        # valida IP singolo
        try:
            ipaddress.IPv4Address(ip)
        except Exception:
            continue
        servizio_to_ips.setdefault(srv, [])
        if ip not in servizio_to_ips[srv]:
            servizio_to_ips[srv].append(ip)

    return servizio_to_ips, location_to_subnets


# ========================================================================
# ROW FILTERING FUNCTIONS
# ========================================================================

def should_include_row(row: dict, wave_filter: str, server_filter: list, 
                       filter_columns: tuple, case_sensitive: bool) -> bool:
    """
    Determine if a row should be included based on all filter criteria.
    
    Args:
        row: CSV row as dictionary
        wave_filter: Wave/group name to filter for
        server_filter: List of server names to filter for
        filter_columns: Column names to check for wave filter
        case_sensitive: Whether wave filter is case-sensitive
        
    Returns:
        True if row passes all filters, False otherwise
    """
    # Filter 1: Check wave/group membership
    if wave_filter:
        if case_sensitive:
            wave_match = any(wave_filter in (row.get(col) or "") for col in filter_columns)
        else:
            wave_lower = wave_filter.lower()
            wave_match = any(wave_lower in (row.get(col) or "").lower() for col in filter_columns)
        
        if not wave_match:
            return False
    
    # Filter 2: Check server name filter (if specified)
    if server_filter:
        src_name = (row.get("src_name") or "").strip().lower()
        dest_name = (row.get("dest_name") or "").strip().lower()
        server_match = any(srv.lower() in src_name or srv.lower() in dest_name 
                          for srv in server_filter)
        if not server_match:
            return False
    
    # Filter 3: Exclude self-talking (same source and destination)
    src_name = (row.get("src_name") or "").strip().lower()
    dest_name = (row.get("dest_name") or "").strip().lower()
    if src_name and dest_name and src_name == dest_name:
        return False
    
    return True


# ========================================================================
# ROW ENRICHMENT FUNCTIONS
# ========================================================================

def enrich_row_with_services(row: dict, ip_to_service: dict) -> None:
    """
    Add service information based on source and destination IPs.
    
    Args:
        row: CSV row to modify (modified in-place)
        ip_to_service: Dictionary mapping IPs to service names
    """
    src_ip = (row.get("src_addr") or "").strip()
    dest_ip = (row.get("dest_addr") or "").strip()
    
    row["src_service"] = ip_to_service.get(src_ip, "unknown")
    row["dest_service"] = ip_to_service.get(dest_ip, "unknown")


def enrich_row_with_locations(row: dict, subnet_index: list) -> None:
    """
    Add location information based on source and destination IPs.
    
    Args:
        row: CSV row to modify (modified in-place)
        subnet_index: Sorted list of (network, location) tuples
    """
    src_ip = (row.get("src_addr") or "").strip()
    dest_ip = (row.get("dest_addr") or "").strip()
    
    # Find locations via subnet matching
    src_loc = find_location_for_ip(src_ip, subnet_index)
    if src_loc is not None:
        row["src_loc"] = src_loc
    
    dest_loc = find_location_for_ip(dest_ip, subnet_index)
    if dest_loc is not None:
        row["dest_loc"] = dest_loc
    
    # Normalize special location values
    for col in ("src_loc", "dest_loc"):
        current = (row.get(col) or "").strip()
        if current:
            normalized = LOC_NORMALIZATION.get(current.lower())
            if normalized is not None:
                row[col] = normalized


def enrich_row_with_comment(row: dict) -> None:
    """
    Add comment field based on service and port rules.
    
    Args:
        row: CSV row to modify (modified in-place)
    """
    comment = ""
    
    # Rule 1: Skip if either service is known (not unknown)
    src_service = (row.get("src_service") or "").strip().lower()
    dest_service = (row.get("dest_service") or "").strip().lower()
    
    if src_service != "unknown" or dest_service != "unknown":
        comment = "skip shared services"
    else:
        # Rule 2: Check for specific ports
        dest_port = row.get("dest_port")
        for port, label in PORT_COMMENTS.items():
            if port_matches(port, dest_port):
                comment = label
                break
    
    # Only set comment if we have something to say
    if comment:
        row["COMMENTO"] = comment


# ========================================================================
# MAIN PROCESSING LOGIC
# ========================================================================

def run_user_python_code(selected_file: str, selected_excel: str, server_filter: list, wave_filter: str, 
                         log_put, progress_set) -> None:
    """
    Main processing function for network analysis.
    
    Performs filtering, enrichment, and transformation of network traffic CSV data.
    
    Args:
        selected_file: Path to input CSV file
        server_filter: List of server names to filter (optional)
        wave_filter: Wave/group name to filter for (optional)
        log_put: Function to output log messages
        progress_set: Function to update progress bar (0-100)
    """
    
    # Progress messages
    STEPS = [
        "Sto aprendo il barattolo di Nutella…",
        "Prendo una bella cucchiaiata…",
        "Sto spalmando con generosità la Nutella…",
        "Un'altra bella cucchiaiata di Nutella…",
        "Barattolo di Nutella finito!",
    ]
    THRESHOLDS = [0, 25, 50, 75, 100]
    
    # Validation
    if not selected_file:
        raise ValueError("Select a CSV first.")
    
    if not selected_excel:
        raise ValueError("Devi selezionare l'Excel con i fogli 'servizi' e 'location'.")

    log_put(f"Carico mapping da Excel: {selected_excel}")
    srv_map, loc_map = load_mappings_from_excel(selected_excel)

    # Popola i global usando gli stessi nomi (niente fallback)
    global SERVIZIO_TO_IPS, LOCATION_TO_SUBNETS
    SERVIZIO_TO_IPS = srv_map
    LOCATION_TO_SUBNETS = loc_map

    log_put(f" - Servizi caricati: {len(SERVIZIO_TO_IPS)}")
    log_put(f" - Location caricate: {len(LOCATION_TO_SUBNETS)}")

    # Count total rows for progress tracking
    with open(selected_file, "r", newline="", encoding="utf-8") as f:
        total_rows_raw = sum(1 for _ in f) - 1  # minus header
    total_rows_raw = max(total_rows_raw, 0)
    
    # Log initial info
    log_put(f"Found {total_rows_raw:,} rows. Starting…")
    if server_filter:
        log_put(f"Filtering for servers: {', '.join(server_filter)}")
    if wave_filter:
        log_put(f"Filtering for wave: {wave_filter}")
    
    # Build lookup indices (once, before processing)
    ip_to_service = build_ip_to_service_index(SERVIZIO_TO_IPS)
    subnet_index = build_subnet_index(LOCATION_TO_SUBNETS)
    
    # Prepare output path
    out_dir = Path(selected_file).parent
    base_name = "nutella"
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    final_path = out_dir / f"{base_name}_{timestamp}.csv"
    
    # Processing state
    processed = 0
    next_step = 0
    last_ui_progress = 0.0
    last_ui_log = 0.0
    t0 = time.time()
    
    # Filter configuration
    FILTER_COLUMNS = ("src_group", "dest_group")
    CASE_SENSITIVE = True
    
    # ========================================================================
    # SINGLE-PASS PROCESSING: Filter + Enrich + Write
    # ========================================================================
    
    log_put("Conteggio righe da processare...")

    # Counters for server traffic direction
    outbound_count = {srv: 0 for srv in server_filter} if server_filter else {}
    inbound_count = {srv: 0 for srv in server_filter} if server_filter else {}

    with open(selected_file, "r", newline="", encoding="utf-8") as src, \
         NamedTemporaryFile("w", newline="", delete=False, encoding="utf-8", 
                           dir=str(out_dir)) as tmp:
        
        reader = csv.DictReader(src)
        
        # Validate required columns exist
        missing = [c for c in FILTER_COLUMNS if c not in reader.fieldnames]
        if missing:
            raise ValueError(f"Missing columns: {', '.join(missing)}")
        
        if server_filter and ("src_name" not in reader.fieldnames or 
                             "dest_name" not in reader.fieldnames):
            raise ValueError("Colonne 'src_name' o 'dest_name' non trovate nel CSV")
        
        # Prepare output columns (add new enrichment columns)
        fieldnames = list(reader.fieldnames or [])
        for col in ("src_service", "dest_service", "src_loc", "dest_loc", "COMMENTO"):
            if col not in fieldnames:
                fieldnames.append(col)
        
        writer = csv.DictWriter(tmp, fieldnames=fieldnames)
        writer.writeheader()
        
        # Process each row: filter -> enrich -> write
                
        # --- PASSATA 1: conteggio righe che passeranno il filtro ---
        total_rows_to_process = 0
        for row in reader:
            if should_include_row(row, wave_filter, server_filter, FILTER_COLUMNS, CASE_SENSITIVE):
                total_rows_to_process += 1

        log_put(f"Righe da processare: {total_rows_to_process}")
        src.seek(0)
        reader = csv.DictReader(src)



        for row in reader:
            # Check all filter criteria
            if not should_include_row(row, wave_filter, server_filter, 
                                     FILTER_COLUMNS, CASE_SENSITIVE):
                continue
            
            # aggiornamento contatori per ogni riga che passa il filtro
            if server_filter:
                src_name = (row.get("src_name") or "").strip().lower()
                dest_name = (row.get("dest_name") or "").strip().lower()
                for srv in server_filter:
                    srv_lower = srv.lower()
                    # Count outbound: server is source
                    if srv_lower in src_name:
                        outbound_count[srv] += 1
                    # Count inbound: server is destination
                    if srv_lower in dest_name:
                        inbound_count[srv] += 1

            # Enrich row with additional data
            enrich_row_with_services(row, ip_to_service)
            enrich_row_with_locations(row, subnet_index)
            enrich_row_with_comment(row)
            
            # Write enriched row
            writer.writerow(row)
            
            # Update progress tracking
            processed += 1
            now = time.time()
            pct = int(processed * 100 / total_rows_to_process) if total_rows_to_process else 0
            
            # Show step messages at thresholds
            if next_step < len(THRESHOLDS) and pct >= THRESHOLDS[next_step]:
                log_put(STEPS[next_step])
                next_step += 1
            
            # Throttle UI updates to avoid overhead
            if now - last_ui_progress >= 1 or processed == total_rows_to_process:
                progress_set(pct)
                last_ui_progress = now
            
            if now - last_ui_log >= 2 or processed == total_rows_to_process:
                log_put(f"Row {processed}/{total_rows_to_process}  ({pct}% done)")
                last_ui_log = now
        
        tmp_path = tmp.name
    
    # Move temp file to final destination
    progress_set(100)
    os.replace(tmp_path, final_path)
    
    log_put(f"Fatto! File processato: {os.path.basename(selected_file)}")
    log_put(f"File generato: {os.path.basename(final_path)}")

    # stampa riepilogo contatori
    if server_filter:
        log_put("---- Riepilogo conteggi per server ----")
        for srv in server_filter:
            out_v = outbound_count.get(srv, 0)
            in_v = inbound_count.get(srv, 0)
            log_put(f"{srv}: outbound_count={out_v} | inbound_count={in_v}")


# ========================================================================
# GUI COMPONENTS  —  “Ferrero / Nutella Edition”
# ========================================================================

class QuizWindow(tk.Toplevel):
    """Interactive quiz window for user engagement."""
    def __init__(self, parent, questions, on_finish=None, title="Quiz Nutella Edition"):
        super().__init__(parent)
        self.parent = parent
        self.questions = questions
        self.on_finish = on_finish
        self.index = 0
        self.answers = []

        # Palette Ferrero / Nutella
        BG = "#f8f4f0"        # panna
        PANEL = "#fffdfb"     # chiaro
        CHOC = "#4b2e05"      # cioccolato
        ACCENT = "#d12727"    # rosso nutella
        ACCENT_HL = "#ff4c4c"
        TEXT = "#3b2a14"

        self.configure(bg=BG)
        self.title(title)
        self.geometry("520x360")
        self.minsize(460, 320)

        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure("Quiz.Card.TFrame", background=PANEL)
        style.configure("Quiz.H2.TLabel", font=("Segoe UI Semibold", 13),
                        background=PANEL, foreground=TEXT)
        style.configure("Quiz.Muted.TLabel", background=PANEL, foreground="#6b5b46")
        style.configure("Quiz.Pill.TButton", padding=(14, 8))
        style.map("Quiz.Pill.TButton",
                  background=[("!disabled", "#f2e6da"), ("active", "#eedbcc")],
                  foreground=[("!disabled", TEXT)])
        style.configure("Quiz.Accent.TButton", background=ACCENT, foreground="white", padding=10)
        style.map("Quiz.Accent.TButton",
                  background=[("!disabled", ACCENT), ("active", ACCENT_HL)])

        self.frame = ttk.Frame(self, style="Quiz.Card.TFrame", padding=20)
        self.frame.pack(fill="both", expand=True)

        self.lbl_q = ttk.Label(self.frame, text="", style="Quiz.H2.TLabel",
                               wraplength=460, justify="left")
        self.lbl_q.pack(anchor="w", pady=(0, 12))

        self.opts_frame = ttk.Frame(self.frame, style="Quiz.Card.TFrame")
        self.opts_frame.pack(fill="both", expand=True)

        self.lbl_step = ttk.Label(self.frame, text="", style="Quiz.Muted.TLabel")
        self.lbl_step.pack(anchor="e", pady=(8, 0))

        self._render()

    def _render(self):
        for w in self.opts_frame.winfo_children():
            w.destroy()

        if self.index >= len(self.questions):
            self._finish()
            return

        q = self.questions[self.index]
        self.lbl_q.configure(text=q["q"])
        self.lbl_step.configure(text=f"{self.index + 1} / {len(self.questions)}")

        for opt in q["options"]:
            btn = ttk.Button(
                self.opts_frame,
                text=opt,
                style="Quiz.Pill.TButton",
                command=lambda opt_text=opt: self._choose(opt_text),
            )
            btn.pack(anchor="w", fill="x", pady=4)

    def _choose(self, opt_text):
        self.answers.append((self.index, opt_text))
        self.index += 1
        self._render()

    def _finish(self):
        if callable(self.on_finish):
            self.on_finish(self.answers)
        messagebox.showinfo("Fatto!", "Grazie! Hai completato il quiz 🍫")
        self.destroy()


class App(tk.Tk):
    """Main application window — Ferrero / Nutella Edition."""
    def __init__(self):
        super().__init__()
        self.title("Network Analysis — Nutella Edition by Angelo Martino")
        self.geometry("1080x740")
        self.minsize(880, 800)

        # Palette Ferrero
        self.COL_BG = "#f5f2ef"       # panna
        self.COL_CARD = "#fffdfa"     # bianco crema
        self.COL_SURF = "#f0e7df"     # beige caldo
        self.COL_TEXT = "#3b2a14"     # marrone scuro
        self.COL_MUTED = "#7b6955"    # marrone chiaro
        self.COL_ACCENT = "#d12727"   # rosso nutella
        self.COL_ACCENT_HL = "#ff4c4c"
        self.COL_BORDER = "#dac7b5"
        self.COL_BTN = "#f7ede3"

        self.configure(bg=self.COL_BG)
        self.selected_file = tk.StringVar(value="")
        self.selected_excel = tk.StringVar(value="")
        self.server_filter = tk.StringVar(value="")
        self.wave_filter = tk.StringVar(value="Wave4")
        self.is_running = tk.BooleanVar(value=False)
        self.ui_queue = queue.Queue()

        self._setup_styles()
        self._create_widgets()
        self.after(60, self._drain_queue)

    def _setup_styles(self):
        s = ttk.Style()
        try:
            s.theme_use("clam")
        except tk.TclError:
            pass

        s.configure("App.TFrame", background=self.COL_BG)
        s.configure("Card.TFrame", background=self.COL_CARD, bordercolor=self.COL_BORDER)
        s.configure("Surface.TFrame", background=self.COL_SURF)
        s.configure("H1.TLabel", font=("Segoe UI Semibold", 18),
                    background=self.COL_BG, foreground=self.COL_TEXT)
        s.configure("H2.TLabel", font=("Segoe UI Semibold", 13),
                    background=self.COL_CARD, foreground=self.COL_TEXT)
        s.configure("Muted.TLabel", background=self.COL_CARD, foreground=self.COL_MUTED)
        s.configure("TEntry", fieldbackground="#fff8f3", foreground=self.COL_TEXT,
                    bordercolor=self.COL_BORDER, insertcolor=self.COL_TEXT)
        s.configure("TButton", font=("Segoe UI", 10), padding=8, background=self.COL_BTN)
        s.map("TButton",
              background=[("!disabled", self.COL_BTN), ("active", "#f3e3d8")])
        s.configure("Accent.TButton", background=self.COL_ACCENT, foreground="white")
        s.map("Accent.TButton",
              background=[("!disabled", self.COL_ACCENT), ("active", self.COL_ACCENT_HL)])
        s.configure("TProgressbar", troughcolor="#e8dfd5", background=self.COL_ACCENT)

    def _create_widgets(self):
        outer = ttk.Frame(self, style="App.TFrame", padding=16)
        outer.pack(fill="both", expand=True)

        # Header
        topbar = ttk.Frame(outer, style="App.TFrame")
        topbar.pack(fill="x", pady=(0, 12))
        ttk.Label(topbar, text="Analisi Network 🍫", style="H1.TLabel").pack(side="left")
        self.status_chip = ttk.Label(topbar, text="Pronto", style="H2.TLabel")
        self.status_chip.pack(side="right")

        # Layout split
        content = ttk.Frame(outer, style="App.TFrame")
        content.pack(fill="both", expand=True)
        content.grid_columnconfigure(0, weight=4, uniform="cols")
        content.grid_columnconfigure(1, weight=6, uniform="cols")
        content.grid_rowconfigure(0, weight=1)

        self.left = ttk.Frame(content, style="Card.TFrame", padding=16)
        self.left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        self.right = ttk.Frame(content, style="Card.TFrame", padding=16)
        self.right.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        self._create_left_panel()
        self._create_right_panel()

        footer = ttk.Frame(outer, style="App.TFrame")
        footer.pack(fill="x", pady=(12, 0))
        ttk.Label(footer, text="© 2025 — Network Analysis Managed by Angelo Martino",
                  style="Muted.TLabel").pack(side="left")
        self.progress = ttk.Progressbar(footer, orient="horizontal",
                                        mode="determinate", length=220)
        self.progress.pack(side="right")

    def _create_left_panel(self):
        ttk.Label(self.left, text="Seleziona report Flexera (.csv)", style="H2.TLabel").pack(anchor="w")
        fr = ttk.Frame(self.left, style="Card.TFrame")
        fr.pack(fill="x", pady=4)
        ttk.Entry(fr, textvariable=self.selected_file).pack(side="left", fill="x", expand=True)
        ttk.Button(fr, text="Sfoglia…", command=self._on_browse).pack(side="left", padx=(8, 0))

        ttk.Label(self.left, text="Seleziona mapping da Excel (.xlsx)", style="H2.TLabel").pack(anchor="w", pady=(12, 2))
        frx = ttk.Frame(self.left, style="Card.TFrame")
        frx.pack(fill="x", pady=4)
        ttk.Entry(frx, textvariable=self.selected_excel).pack(side="left", fill="x", expand=True)
        ttk.Button(frx, text="Sfoglia…", command=self._on_browse_excel).pack(side="left", padx=(8, 0))

        ttk.Label(self.left, text="Filtra Wave (es: Wave4)", style="H2.TLabel").pack(anchor="w", pady=(12, 2))
        ttk.Entry(self.left, textvariable=self.wave_filter).pack(fill="x", pady=(0, 4))
        ttk.Label(self.left, text="Filtra server (es: SERVER1, SERVER2)", style="H2.TLabel").pack(anchor="w", pady=(12, 2))
        ttk.Entry(self.left, textvariable=self.server_filter).pack(fill="x")

        btns = ttk.Frame(self.left, style="Card.TFrame")
        btns.pack(fill="x", pady=16)
        self.run_btn = ttk.Button(btns, text="Esegui", style="Accent.TButton", command=self._on_run)
        self.run_btn.pack(side="left")
        ttk.Button(btns, text="Captcha", command=self._open_captcha).pack(side="left", padx=8)

        # Immagine anteprima
        self._create_preview_image(self.left)
    
    def _create_preview_image(self, parent):
        block = ttk.Frame(parent, style="Card.TFrame")
        block.pack(fill="both", expand=True, pady=(10, 0))
        ttk.Label(block, text="Risultato atteso:", style="H2.TLabel").pack(anchor="w")

        self.image_container = ttk.Frame(block, style="Surface.TFrame")
        self.image_container.pack(fill="both", expand=True, pady=(8, 0))
        self.image_container.configure(width=400, height=160)
        self.image_container.pack_propagate(False)

        self._orig_bg_pil = None
        self._bg_photo = None
        self.image_label = tk.Label(self.image_container, bd=0,
                                    highlightthickness=0, bg=self.COL_SURF)
        self.image_label.pack(fill="both", expand=True)

        try:
            base_dir = Path(__file__).resolve().parent
        except Exception:
            base_dir = Path.cwd()
        bg_path = base_dir / "background.png"
        if bg_path.exists():
            try:
                self._orig_bg_pil = Image.open(bg_path).convert("RGBA")
            except Exception as e:
                print(f"Could not load image: {e}")

        def _render_preview(event=None):
            w, h = self.image_container.winfo_width(), self.image_container.winfo_height()
            if w < 3 or h < 3:
                return
            canvas = Image.new("RGBA", (w, h), (245, 242, 239, 255))
            if self._orig_bg_pil:
                src = self._orig_bg_pil
                sw, sh = src.size
                ratio = min(w / sw, h / sh)
                nw, nh = int(sw * ratio), int(sh * ratio)
                resized = src.resize((nw, nh), Image.LANCZOS)
                x, y = (w - nw) // 2, (h - nh) // 2
                canvas.paste(resized, (x, y), resized)
            self._bg_photo = ImageTk.PhotoImage(canvas)
            self.image_label.configure(image=self._bg_photo)

        self.image_container.bind("<Configure>", _render_preview)
        self.after(50, _render_preview)

    def _create_right_panel(self):
        ttk.Label(self.right, text="Output", style="H2.TLabel").pack(anchor="w")
        console_frame = ttk.Frame(self.right, style="Surface.TFrame")
        console_frame.pack(fill="both", expand=True, pady=(8, 0))
        yscroll = tk.Scrollbar(console_frame, orient="vertical")
        yscroll.pack(side="right", fill="y")
        self.console = tk.Text(console_frame, height=20, wrap="word",
                               bg="#fffaf5", fg=self.COL_TEXT,
                               insertbackground=self.COL_TEXT,
                               relief="flat", padx=12, pady=10,
                               font=("Consolas", 10))
        self.console.pack(fill="both", expand=True)
        self.console.config(yscrollcommand=yscroll.set)
        yscroll.config(command=self.console.yview)
        ttk.Button(self.right, text="Pulisci Output", command=self._clear_output).pack(pady=(10, 0))

    # === Eventi standard (invariati) ===
    def _on_browse(self):
        path = filedialog.askopenfilename(title="Seleziona CSV", filetypes=[("CSV files", "*.csv"), ("Tutti i file", "*.*")])
        if path:
            self.selected_file.set(path)
            self._log(f"Selezionato: {path}")

    def _on_browse_excel(self):
        path = filedialog.askopenfilename(
            title="Seleziona Excel con 'servizi' e 'location'",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("Tutti i file", "*.*")]
        )
        if path:
            self.selected_excel.set(path)
            self._log(f"Excel mapping selezionato: {path}")

    def _on_run(self):
        """Gestisce il click sul bottone 'Esegui'."""
        if self.is_running.get():
            return

        # --- Validazioni preliminari ---
        if not self.selected_file.get().strip():
            messagebox.showerror("Errore", "Devi selezionare il file CSV (report Flexera).")
            return

        if not self.selected_excel.get().strip():
            messagebox.showerror("Errore", "Devi selezionare il file Excel con i fogli 'servizi' e 'location'.")
            # 🔧 FIX: assicurati che il bottone torni cliccabile
            self.is_running.set(False)
            self.run_btn.state(["!disabled"])
            self._set_status("Pronto")
            return

        # --- Se tutto ok, prosegui ---
        self.is_running.set(True)
        self.run_btn.state(["disabled"])
        self._set_status("Elaborazione…")
        self.progress["value"] = 0
        self._clear_output()

        server_text = self.server_filter.get().strip()
        server_filter = [s.strip() for s in server_text.split(",") if s.strip()] if server_text else []
        wave_filter = self.wave_filter.get().strip()

        t = threading.Thread(
            target=self._worker_wrapper,
            args=(
                self.selected_file.get(),
                self.selected_excel.get(),
                server_filter,
                wave_filter,
            ),
            daemon=True,
        )
        t.start()

    def _worker_wrapper(self, selected_file, selected_excel, server_filter, wave_filter):
        def queue_put(msg): self.ui_queue.put(("log", msg))
        def progress_set(val): self.ui_queue.put(("progress", val))
        try:
            run_user_python_code(selected_file, selected_excel, server_filter, wave_filter, queue_put, progress_set)
            self.ui_queue.put(("status", "Completato"))
        except Exception as e:
            self.ui_queue.put(("log", f"[ERRORE] {e}"))
            self.ui_queue.put(("status", "Fallito"))
        finally:
            self.ui_queue.put(("done", None))


    def _open_captcha(self):
        questions = [
            {"q": "Quale sede ha finito prima la Nutella?",
             "options": ["Alba", "Torino", "Milano", "Lussemburgo"]},
            {"q": "Qual è il miglior snack Ferrero?",
             "options": ["Kinder Bueno", "Nutella B-Ready", "Ferrero Rocher", "Tutti"]},
        ]
        def on_finish(answers):
            self._log("Quiz completato!")
            for idx, opt in answers:
                self._log(f"  Q{idx+1}: {opt}")
        QuizWindow(self, questions, on_finish=on_finish)

    def _clear_output(self):
        self.console.delete("1.0", "end")

    def _drain_queue(self):
        try:
            while True:
                key, payload = self.ui_queue.get_nowait()
                if key == "log":
                    self._log(payload)
                elif key == "progress":
                    self.progress["value"] = payload
                elif key == "status":
                    self._set_status(payload)
                elif key == "done":
                    self.is_running.set(False)
                    self.run_btn.state(["!disabled"])
        except queue.Empty:
            pass
        finally:
            self.after(60, self._drain_queue)

    def _log(self, msg):
        self.console.insert("end", msg + "\n")
        self.console.see("end")

    def _set_status(self, text):
        self.status_chip.configure(text=text)


if __name__ == "__main__":
    app = App()
    app.mainloop()
